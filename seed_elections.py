"""
seed_elections.py
=================
Reads AE_AP.xlsx (4 sheets: 2009, 2014, 2019, 2024) and seeds / updates
Firestore pollingStation documents with election data.

Document ID format: "{ac_id}-{ps_no}"  (e.g., "108-1")
Election fields:    election2009, election2014, election2019, election2024

Schema per field:
{
  "year": 2009,
  "total_votes": 847,
  "candidates": {
    "TDP": 0.423,   <- fraction (0-1)
    "INC": 0.215,
    ...
  }
}

Usage:
  pip install firebase-admin openpyxl
  # Place serviceAccount.json in same folder
  python seed_elections.py
"""

import json
import os
import sys
import openpyxl

try:
    import firebase_admin
    from firebase_admin import credentials, firestore
except ImportError:
    print("Install firebase-admin:  pip install firebase-admin openpyxl")
    sys.exit(1)

# ── CONFIG ─────────────────────────────────────────────────────────────────
SERVICE_ACCOUNT = "jspdboard-fa45e-firebase-adminsdk-fbsvc-9092b52943.json"
XLSX_FILE       = "AE_AP.xlsx"
COLLECTION      = "pollingStations"
BATCH_SIZE      = 400          # Firestore batch limit is 500
# ───────────────────────────────────────────────────────────────────────────

def die(msg):
    print("ERROR:", msg); sys.exit(1)

if not os.path.exists(SERVICE_ACCOUNT):
    die(f"{SERVICE_ACCOUNT} not found.\n"
        "Download from Firebase Console → Project Settings → Service Accounts → Generate new private key")

cred = credentials.Certificate(SERVICE_ACCOUNT)
firebase_admin.initialize_app(cred)
db = firestore.client()
print("✓ Firebase connected")

# ── HELPERS ────────────────────────────────────────────────────────────────

def si(v):
    """Safe int"""
    if v is None: return 0
    try: return int(float(v))
    except: return 0

def sf(v):
    """Safe float"""
    if v is None: return 0.0
    try: return float(v)
    except: return 0.0

def share(votes, total):
    if total > 0: return round(votes / total, 6)
    return 0.0

def election_field(year: int, party_votes: dict, total: int) -> dict:
    candidates = {p: share(v, total) for p, v in party_votes.items() if p and v > 0}
    return {"year": year, "total_votes": total, "candidates": candidates}

def normalize_party(name: str) -> str:
    """Normalize common party name variants."""
    name = str(name).strip().upper()
    aliases = {
        "JANASENA ": "JANASENA",
        "JANA SENA": "JANASENA",
        "JATIYA JANA SENA PARTY": "JANASENA",
        "JSP": "JANASENA",
        "YSRCP": "YSRCP",
        "Y.S.R.C.P": "YSRCP",
        "TDP": "TDP",
        "TELUGUDE SHAM PARTY": "TDP",
        "INC": "INC",
        "CONGRESS": "INC",
        "BAHUJAN SAMAJ PARTY": "BSP",
        "B.S.P": "BSP",
    }
    return aliases.get(name, name)

# ── PARSERS ────────────────────────────────────────────────────────────────

def parse_main_2009(wb):
    """Sheet 'main' — 2009 election data.
    Columns: 0=Year 1=Type 2=District 3=PC 4=AssemblyID 5=Assembly
             6=Town 7=Mandal 8=VillageID 9=Village 10=BoothID 11=Booth#
             12=PRP 13=INC 14=IND 15=TDP 16=LSP 17=BJP …
    """
    ws = wb['main']
    results = {}   # (ac_id, ps_no) -> election_field
    PARTY_COLS = [(12,'PRP'),(13,'INC'),(14,'IND'),(15,'TDP'),(16,'LSP'),(17,'BJP')]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        if si(row[0]) != 2009: continue
        ac_id = str(si(row[4]))
        ps_no = str(si(row[11]))
        if not ac_id or ps_no == '0': continue

        party_votes = {}
        for col, name in PARTY_COLS:
            if col < len(row):
                v = si(row[col])
                if v: party_votes[name] = v
        # Sum any additional columns as OTHERS
        others = 0
        for col in range(18, min(len(row), 30)):
            v = si(row[col])
            if v: others += v
        if others: party_votes['OTHERS'] = others

        total = sum(party_votes.values())
        if total == 0: continue
        results[(ac_id, ps_no)] = election_field(2009, party_votes, total)

    print(f"  2009: parsed {len(results)} booths")
    return results


def parse_sheet1_2014(wb):
    """Sheet 'Sheet1' — 2014 election data.
    Columns: 0-11 same prefix
             12=YSRCP 13=TDP 14=JSP 15=NOTA 16=BSP 17=LSP 18=INC 19=Total
    """
    ws = wb['Sheet1']
    results = {}
    PARTY_COLS = [(12,'YSRCP'),(13,'TDP'),(14,'JANASENA'),(15,'NOTA'),
                  (16,'BSP'),(17,'LSP'),(18,'INC')]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        if si(row[0]) != 2014: continue
        ac_id = str(si(row[4]))
        ps_no = str(si(row[11]))
        if not ac_id or ps_no == '0': continue

        party_votes = {}
        for col, name in PARTY_COLS:
            if col < len(row):
                v = si(row[col])
                if v: party_votes[name] = v

        total = si(row[19]) if len(row) > 19 and row[19] else sum(party_votes.values())
        if total == 0: continue
        results[(ac_id, ps_no)] = election_field(2014, party_votes, total)

    print(f"  2014: parsed {len(results)} booths")
    return results


def parse_sheet2_2019(wb):
    """Sheet 'Sheet2' — 2019 election data.
    Columns: 0-11 same prefix
             12=TDP 13=YSRCP 14=JANASENA 15=BJP 16=INC 17=IND 18=NOTA
    """
    ws = wb['Sheet2']
    results = {}
    PARTY_COLS = [(12,'TDP'),(13,'YSRCP'),(14,'JANASENA'),(15,'BJP'),
                  (16,'INC'),(17,'IND'),(18,'NOTA')]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        if si(row[0]) != 2019: continue
        ac_id = str(si(row[4]))
        ps_no = str(si(row[11]))
        if not ac_id or ps_no == '0': continue

        party_votes = {}
        for col, name in PARTY_COLS:
            if col < len(row):
                v = si(row[col])
                if v: party_votes[name] = v

        total = sum(party_votes.values())
        if total == 0: continue
        results[(ac_id, ps_no)] = election_field(2019, party_votes, total)

    print(f"  2019: parsed {len(results)} booths")
    return results


def parse_sheet4_2024(wb):
    """Sheet 'Sheet4' — 2024 election data.
    Different format: groups by assembly name, then headers row, then booth rows.
    Row format: Assembly|PS No.|Party1|Party2|...|Total|NOTA
    
    Since we don't have ac_id in this sheet, we match by ac_name lookup later.
    Returns dict: assembly_name.upper() -> {ps_no -> election_field}
    """
    ws = wb['Sheet4']
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))

    # Build assembly name -> ac_id map from assemblies.ts (we'll pass it in)
    result = {}  # assembly_name.upper() -> {ps_no_str -> election_field}

    i = 0
    while i < len(all_rows):
        row = all_rows[i]
        # Assembly name marker: col[0] = name string, col[1] = None
        if (row[0] and isinstance(row[0], str) and
                str(row[0]).strip() and
                (len(row) < 2 or row[1] is None)):
            asm_name = str(row[0]).strip().upper()
            i += 1
            if i >= len(all_rows): break

            # Next row = headers
            hdr_row = all_rows[i]
            i += 1

            # Parse header columns
            headers = [str(h).strip() if h else '' for h in hdr_row]

            # Find PS No column (should be 'PS No.' or similar at index 1)
            ps_col = 1
            # Find NOTA col
            nota_col = None
            for idx, h in enumerate(headers):
                if 'NOTA' in h.upper():
                    nota_col = idx
            # Find Total col
            total_col = None
            for idx, h in enumerate(headers):
                if 'TOTAL' in h.upper():
                    total_col = idx

            # Party columns: indices 2..N (skip 0=asm, 1=ps_no, total, nota, rejected)
            skip_terms = {'TOTAL', 'NOTA', 'NO.OF', 'REJECTED', 'PS NO', ''}
            party_cols = []
            for idx, h in enumerate(headers):
                if idx in (0, 1): continue
                hu = h.upper()
                if any(s in hu for s in skip_terms): continue
                if h: party_cols.append((idx, normalize_party(h)))

            # Read booth rows until next assembly marker or end
            booth_data = {}
            while i < len(all_rows):
                dr = all_rows[i]
                # Next assembly block?
                if (dr[0] and isinstance(dr[0], str) and
                        str(dr[0]).strip() and
                        (len(dr) < 2 or dr[1] is None)):
                    break
                ps_num = dr[1] if len(dr) > 1 else None
                if ps_num is None or not isinstance(ps_num, (int, float)):
                    i += 1; continue
                ps_no = str(si(ps_num))

                party_votes = {}
                for col_idx, pname in party_cols:
                    if col_idx < len(dr):
                        v = si(dr[col_idx])
                        if v: party_votes[pname] = v
                # NOTA
                if nota_col is not None and nota_col < len(dr):
                    v = si(dr[nota_col])
                    if v: party_votes['NOTA'] = v

                total = si(dr[total_col]) if total_col and total_col < len(dr) else sum(party_votes.values())
                if total == 0:
                    i += 1; continue

                booth_data[ps_no] = election_field(2024, party_votes, total)
                i += 1

            if booth_data:
                result[asm_name] = booth_data
        else:
            i += 1

    total_booths = sum(len(v) for v in result.values())
    print(f"  2024: parsed {total_booths} booths across {len(result)} assemblies")
    return result


# ── MAIN SEEDER ────────────────────────────────────────────────────────────

def build_ac_name_map():
    """Returns dict: AC_NAME.upper() -> ac_id (str) from Firebase assemblies collection."""
    # We'll derive from Firestore or use a local hardcoded list.
    # First try from pollingStations to discover ac_id <-> ac_name mappings.
    print("  Building assembly name→ID map from Firestore...")
    mapping = {}
    try:
        for doc in db.collection(COLLECTION).stream():
            d = doc.to_dict()
            ac_id = str(d.get('ac_id', ''))
            ac_name = str(d.get('ac_name', '')).upper()
            if ac_id and ac_name:
                mapping[ac_name] = ac_id
    except Exception as e:
        print(f"  Warning: {e}")
    print(f"  Found {len(mapping)} unique assemblies in DB")
    return mapping


def seed():
    print(f"\nLoading {XLSX_FILE}...")
    wb = openpyxl.load_workbook(XLSX_FILE, data_only=True, read_only=True)

    print("\nParsing sheets...")
    data_2009 = parse_main_2009(wb)
    data_2014 = parse_sheet1_2014(wb)
    data_2019 = parse_sheet2_2019(wb)
    data_2024_by_name = parse_sheet4_2024(wb)
    wb.close()

    # ── Merge 2009/2014/2019 by (ac_id, ps_no) key ─────────────────────
    all_keys = set(data_2009.keys()) | set(data_2014.keys()) | set(data_2019.keys())
    print(f"\nTotal unique (ac_id, ps_no) combinations: {len(all_keys)}")

    # Prepare batch writes
    total_written = 0

    def flush(batch, count):
        nonlocal total_written
        if count:
            batch.commit()
            total_written += count
            print(f"  ↳ Committed {count} docs (total: {total_written})")
        return db.batch(), 0

    # ── Write 2009, 2014, 2019 ──────────────────────────────────────────
    print("\nUploading 2009 / 2014 / 2019 election data...")
    batch = db.batch()
    bcount = 0

    for ac_id, ps_no in sorted(all_keys):
        doc_id = f"{ac_id}-{ps_no}"
        ref = db.collection(COLLECTION).document(doc_id)

        update = {}
        if (ac_id, ps_no) in data_2009:
            update['election2009'] = data_2009[(ac_id, ps_no)]
        if (ac_id, ps_no) in data_2014:
            update['election2014'] = data_2014[(ac_id, ps_no)]
        if (ac_id, ps_no) in data_2019:
            update['election2019'] = data_2019[(ac_id, ps_no)]

        if not update:
            continue

        # set with merge so we don't wipe existing fields
        batch.set(ref, update, merge=True)
        bcount += 1

        if bcount >= BATCH_SIZE:
            batch, bcount = flush(batch, bcount)

    batch, bcount = flush(batch, bcount)

    # ── Write 2024 ──────────────────────────────────────────────────────
    if data_2024_by_name:
        print("\nUploading 2024 election data...")
        ac_name_map = build_ac_name_map()

        batch = db.batch()
        bcount = 0
        unmatched_asms = []

        for asm_name, booth_map in data_2024_by_name.items():
            ac_id = ac_name_map.get(asm_name)
            if not ac_id:
                # Try partial match
                for db_name, db_id in ac_name_map.items():
                    if asm_name in db_name or db_name in asm_name:
                        ac_id = db_id
                        break

            if not ac_id:
                unmatched_asms.append(asm_name)
                continue

            for ps_no, ef in booth_map.items():
                doc_id = f"{ac_id}-{ps_no}"
                ref = db.collection(COLLECTION).document(doc_id)
                batch.set(ref, {"election2024": ef}, merge=True)
                bcount += 1
                if bcount >= BATCH_SIZE:
                    batch, bcount = flush(batch, bcount)

        batch, bcount = flush(batch, bcount)

        if unmatched_asms:
            print(f"\n  ⚠ Could not match {len(unmatched_asms)} assemblies from 2024 sheet:")
            for a in unmatched_asms[:10]:
                print(f"    - {a}")

    # ── Save 2024 as JSON backup ─────────────────────────────────────────
    out_path = "election2024_parsed.json"
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(data_2024_by_name, f, indent=2)
    print(f"\n  2024 data also saved to {out_path} as backup")

    print(f"\n{'='*50}")
    print(f"✓ DONE — Total Firestore writes: {total_written}")
    print(f"{'='*50}")


if __name__ == '__main__':
    seed()
